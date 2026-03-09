import os
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import Paragraph
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_LEFT
from datetime import datetime
from num2words import num2words

def valor_por_extenso(valor):
    """Converte valor para extenso"""
    try:
        # Separa reais e centavos
        reais = int(valor)
        centavos = int((valor - reais) * 100)
        
        extenso_reais = num2words(reais, lang='pt_BR')
        
        if centavos > 0:
            extenso_centavos = num2words(centavos, lang='pt_BR')
            if reais == 1:
                return f"{extenso_reais} real e {extenso_centavos} centavos"
            else:
                return f"{extenso_reais} reais e {extenso_centavos} centavos"
        else:
            if reais == 1:
                return f"{extenso_reais} real"
            else:
                return f"{extenso_reais} reais"
    except:
        return f"{valor:.2f} reais"

def desenhar_conteudo_recibo(c, pessoa, width, height, margin):
    """Função auxiliar para desenhar o conteúdo do recibo com cabeçalho oficial"""
    
    # --- Cabeçalho Oficial ---
    y_header = height - 3.5 * cm
    
    # 1. Brasão (Tenta carregar se existir)
    # Altere o caminho abaixo para o local real da imagem no seu servidor
    caminho_brasao = os.path.join(os.path.dirname(__file__), '..', 'static', 'images', 'brasao.jpg')
    if os.path.exists(caminho_brasao):
        c.drawImage(caminho_brasao, (width/2) - 1.0*cm, y_header, width=1.5*cm, height=1.5*cm, mask='auto')
        y_header -= 1.5 * cm
    else:
        # Se não houver imagem, apenas pula o espaço
        y_header -= 1.5 * cm
        
    # 2. Textos do Cabeçalho
    c.setFont("Helvetica-Bold", 12)
    c.drawCentredString(width / 2, y_header, "ESTADO DA PARAÍBA")
    y_header -= 0.5 * cm
    c.drawCentredString(width / 2, y_header, "PREFEITURA MUNICIPAL DE POCINHOS")
    y_header -= 0.5 * cm
    c.drawCentredString(width / 2, y_header, "SECRETARIA MUNICIPAL DE ASSISTÊNCIA DE SOCIAL")
    
    # --- Título "RECIBO" ---
    y_recibo = y_header - 2.5 * cm
    c.setFont("Helvetica-Bold", 25)
    c.drawCentredString(width / 2, y_recibo, "R E C I B O")
    
    # --- Valor ---
    c.setFont("Helvetica-Bold", 11)
    valor_formatado = f"{pessoa.valor_beneficio:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    c.drawRightString(width - margin, y_recibo - 2 * cm, f"Valor R$ {valor_formatado}")
    
    # --- Configuração do Texto Justificado com Parágrafo ---
    styles = getSampleStyleSheet()
    style_justificado = ParagraphStyle(
        'CustomJustify',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=12,
        leading=18,
        alignment=TA_JUSTIFY,
        firstLineIndent=2 * cm
    )
    
    meses = [
        "JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
        "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"
    ]
    mes_nome = meses[datetime.now().month - 1]
    valor_ext = valor_por_extenso(pessoa.valor_beneficio).upper()
    
    nome_beneficio = pessoa.beneficio.nome if hasattr(pessoa, 'beneficio') else 'Auxílio'
    texto_html = (
        f"Recebi da Prefeitura Municipal de Pocinhos, a importância de "
        f"<b>R$ {valor_formatado} ({valor_ext})</b>, em virtude da necessidade de que "
        f"seja garantido o pagamento referente ao {nome_beneficio}, "
        f"conforme PARECER SOCIAL FAVORÁVEL da Assistente Social do CRAS, "
        f"referente ao mês de <b>{mes_nome}</b> do corrente ano."
    )
    
    p1 = Paragraph(texto_html, style_justificado)
    largura_util = width - (2 * margin)
    w, h = p1.wrap(largura_util, height)
    y_pos = y_recibo - 4 * cm - h
    p1.drawOn(c, margin, y_pos)
    
    texto_quitacao = "Pelo que emito o presente recibo em duas vias de igual teor, dando-lhe plena e total quitação."
    p2 = Paragraph(texto_quitacao, style_justificado)
    w2, h2 = p2.wrap(largura_util, height)
    y_pos = y_pos - h2 - 0.5 * cm
    p2.drawOn(c, margin, y_pos)
    
    # --- Local e Data ---
    y_pos -= 2 * cm
    c.setFont("Helvetica", 12)
    ano_atual = datetime.now().year
    c.drawRightString(width - margin, y_pos, f"Pocinhos, _______/_______/ {ano_atual}")
    
    # --- Linha de Assinatura ---
    y_pos -= 3.5 * cm
    c.setLineWidth(0.5)
    c.line(margin + 1*cm, y_pos, width - margin - 1*cm, y_pos)
    
    # --- Dados do Beneficiário ---
    y_pos -= 0.6 * cm
    c.setFont("Helvetica-Bold", 11)
    nome = getattr(pessoa, 'nome_completo', 'NOME NÃO INFORMADO').upper()
    c.drawCentredString(width / 2, y_pos, nome)
    
    y_pos -= 0.5 * cm
    c.setFont("Helvetica", 10)
    cpf = getattr(pessoa, 'cpf', '000.000.000-00')
    c.drawCentredString(width / 2, y_pos, f"CPF: {cpf}")
    
    endereco = ""
    if hasattr(pessoa, 'endereco'):
        endereco = pessoa.endereco
    elif hasattr(pessoa, 'rua'):
        endereco = f"{pessoa.rua}, {getattr(pessoa, 'numero', '')}"
        
    if endereco:
        y_pos -= 0.5 * cm
        c.drawCentredString(width / 2, y_pos, str(endereco))
    # --- Bairro e Cidade (Adicionado conforme solicitado) ---
    bairro = getattr(pessoa, 'bairro', '')
    cidade = getattr(pessoa, 'cidade', 'Pocinhos/PB') # Padrão Pocinhos se não informado
    
    if bairro or cidade:
        y_pos -= 0.5 * cm
        # Formata como "Bairro - Cidade" ou apenas um deles se o outro faltar
        info_local = f"{bairro} - {cidade}" if bairro and cidade else (bairro or cidade)
        c.drawCentredString(width / 2, y_pos, str(info_local))

def gerar_recibo_paginas_separadas(pessoa):
    """Gera um buffer contendo duas páginas, cada uma com um recibo"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 2.5 * cm
    
    # Primeira Página
    desenhar_conteudo_recibo(c, pessoa, width, height, margin)
    c.showPage()
    
    # Segunda Página
    desenhar_conteudo_recibo(c, pessoa, width, height, margin)
    c.showPage()
    
    c.save()
    buffer.seek(0)
    return buffer
 
def gerar_recibos_massa_pdf(pessoas):
    """Gera recibos em massa (2 vias para cada pessoa)"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 2.5 * cm
    
    for pessoa in pessoas:
        # Primeira via
        desenhar_conteudo_recibo(c, pessoa, width, height, margin)
        c.showPage()
        
        # Segunda via
        desenhar_conteudo_recibo(c, pessoa, width, height, margin)
        c.showPage()
    
    c.save()
    buffer.seek(0)
    return buffer
    
def gerar_documentos_massa_pdf(pessoas):
    """Gera PDF consolidado com documentos (2 vias, máx 4 páginas cada) usando disco"""
    import tempfile
    from PyPDF2 import PdfWriter, PdfReader
    
    # Usa arquivo temporário em disco
    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
        tmp_path = tmp_file.name
    
    writer = PdfWriter()
    arquivos_abertos = []
    
    try:
        for pessoa in pessoas:
            arquivo_path = pessoa.documento.arquivo.path
            
            try:
                pdf_file = open(arquivo_path, 'rb')
                arquivos_abertos.append(pdf_file)
                
                reader = PdfReader(pdf_file)
                total_paginas = min(len(reader.pages), 4)
                
                # 1ª via
                for i in range(total_paginas):
                    writer.add_page(reader.pages[i])
                
                # 2ª via
                for i in range(total_paginas):
                    writer.add_page(reader.pages[i])
                    
            except Exception:
                raise Exception(f"Erro ao processar documento de {pessoa.nome_completo}. Arquivo pode estar corrompido.")
        
        # Escreve no arquivo temporário
        with open(tmp_path, 'wb') as output_file:
            writer.write(output_file)
        
        # Lê o arquivo final para retornar
        with open(tmp_path, 'rb') as final_file:
            buffer = BytesIO(final_file.read())
        
        return buffer
        
    finally:
        # Fecha todos os arquivos abertos
        for f in arquivos_abertos:
            try:
                f.close()
            except:
                pass
        
        # Remove arquivo temporário
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

def gerar_memorando_segunda_via_pdf(memorando):
    """Gera PDF do memorando usando dados do snapshot (segunda via idêntica)"""
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    margin = 2 * cm
    
    # Configurações
    altura_linha = 0.7 * cm
    margem_inferior_tabela = 2 * cm
    espaco_fechamento = 5.5 * cm
    config = memorando
    
    col_widths = [1.0*cm, 6.0*cm, 4.5*cm, 2.5*cm, 3.0*cm]
    total_table_width = sum(col_widths)
    table_margin = (width - total_table_width) / 2
    
    col_x_starts = [table_margin]
    for w in col_widths[:-1]:
        col_x_starts.append(col_x_starts[-1] + w)
    
    styles = getSampleStyleSheet()
    style_intro = ParagraphStyle('Intro', parent=styles['Normal'], fontName='Helvetica', fontSize=11, leading=14, alignment=TA_JUSTIFY)
    
    beneficio = memorando.beneficio
    pessoas = list(memorando.pessoas.all().order_by('ordem'))
    
    def desenhar_cabecalho_pagina():
        y = height - 1.5 * cm
        
        caminho_brasao = os.path.join(os.path.dirname(__file__), '..', 'static', 'images', 'brasao.jpg')
        if os.path.exists(caminho_brasao):
            c.drawImage(caminho_brasao, (width/2) - 0.75*cm, y - 1.5*cm, width=1.5*cm, height=1.5*cm, mask='auto')
        y -= 2.5 * cm
        
        c.setFont("Helvetica-Bold", 11)
        c.setFillColor(colors.black)
        c.drawCentredString(width / 2, y, "ESTADO DA PARAÍBA")
        y -= 0.4 * cm
        c.drawCentredString(width / 2, y, "PREFEITURA MUNICIPAL DE POCINHOS")
        y -= 0.4 * cm
        c.drawCentredString(width / 2, y, "SECRETARIA MUNICIPAL DE ASSISTÊNCIA SOCIAL")
        
        y -= 1.2 * cm
        c.setFont("Helvetica-Bold", 14)
        c.drawCentredString(width / 2, y, f"M E M O R A N D O Nº. {memorando.numero}")
        
        y_quadro_top = y - 0.8 * cm
        quadro_height = 3.2 * cm
        c.setLineWidth(1)
        c.setStrokeColor(colors.navy)
        c.rect(margin, y_quadro_top - quadro_height, width - 2 * margin, quadro_height)
        c.setStrokeColor(colors.black)
        
        y_row1 = y_quadro_top - 0.7 * cm
        y_row2 = y_quadro_top - 1.4 * cm
        y_row3 = y_quadro_top - 2.6 * cm
        
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.navy)
        c.drawString(margin + 0.4 * cm, y_row1, "DE:")
        c.drawString(margin + 0.4 * cm, y_row2, "PARA:")
        c.setFillColor(colors.black)
        c.drawString(margin + 0.4 * cm, y_row3, "ASSUNTO:")
        
        c.setFont("Helvetica", 10)
        c.drawString(margin + 3 * cm, y_row1, "SECRETARIA DE ASSISTÊNCIA SOCIAL")
        
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(colors.navy)
        c.drawString(margin + 3 * cm, y_row2, "SECRETARIA DE FINANÇAS")
        c.setFont("Helvetica", 9)
        c.drawString(margin + 3 * cm, y_row2 - 0.5 * cm, f"Att: Sr. {config.financas_nome} – {config.financas_cargo}")
        
        c.setFillColor(colors.black)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(margin + 3 * cm, y_row3, "Envio de pagamento.")
        
        data_memorando = memorando.created_at.strftime("%d/%m/%Y")
        c.setFont("Helvetica-Bold", 10)
        c.drawRightString(width - margin - 0.4 * cm, y_row1, f"DATA: {data_memorando}")
        
        y_texto = y_quadro_top - quadro_height - 1.0 * cm
        texto_intro = "Por meio do presente, solicito que sejam tomadas as medidas necessárias no sentido de que sejam empenhadas e pagas as despesas que seguem abaixo:"
        p_intro = Paragraph(texto_intro, style_intro)
        w_p, h_p = p_intro.wrap(width - 2 * margin, height)
        y_texto -= h_p
        p_intro.drawOn(c, margin, y_texto)
        
        return y_texto - 1.0 * cm
    
    def desenhar_cabecalho_tabela(y_pos):
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(colors.black)
        header_y_bottom = y_pos - 0.7*cm
        
        c.rect(table_margin, header_y_bottom, total_table_width, 0.7*cm)
        
        current_x = table_margin
        for w in col_widths:
            c.line(current_x, y_pos, current_x, header_y_bottom)
            current_x += w
        c.line(current_x, y_pos, current_x, header_y_bottom)
        
        y_text = y_pos - 0.5*cm
        c.drawCentredString(col_x_starts[0] + col_widths[0]/2, y_text, "N.º")
        c.drawCentredString(col_x_starts[1] + col_widths[1]/2, y_text, "Beneficiário")
        c.drawCentredString(col_x_starts[2] + col_widths[2]/2, y_text, "Tipo de Benefício")
        c.drawCentredString(col_x_starts[3] + col_widths[3]/2, y_text, "Valor")
        c.drawCentredString(col_x_starts[4] + col_widths[4]/2, y_text, "Conta Pagadora")
        
        return header_y_bottom
    
    def desenhar_fechamento_assinatura(y_pos):
        c.setFillColor(colors.black)
        
        texto_fim = "Despeço-me cordialmente aproveitando a oportunidade para reiterar os nossos sentimentos de elevada estima e consideração e me deixando a disposição para o que precisar."
        p_fim = Paragraph(texto_fim, style_intro)
        w_f, h_f = p_fim.wrap(width - 2 * margin, height)
        y_pos -= h_f
        p_fim.drawOn(c, margin, y_pos)
        
        y_pos -= 1.2 * cm
        c.setFont("Helvetica", 11)
        c.drawString(margin, y_pos, "Atenciosamente,")
        
        y_pos -= 2 * cm
        c.setFont("Helvetica-Bold", 13)
        c.drawCentredString(width / 2, y_pos, config.secretaria_nome)
        y_pos -= 0.5 * cm
        c.setFont("Helvetica", 12)
        c.drawCentredString(width / 2, y_pos, config.secretaria_cargo)
        
        y_pos = max(y_pos - 1.5 * cm, 2.5 * cm)
        c.setLineWidth(0.5)
        c.line(margin, y_pos, width - margin, y_pos)
        
        c.setFont("Helvetica", 8)
        c.setFillColor(colors.grey)
        c.drawCentredString(width / 2, y_pos - 0.4 * cm, config.endereco)
        c.drawCentredString(width / 2, y_pos - 0.8 * cm, f"CEP: {config.cep}   –   Pocinhos – PB")
        c.drawCentredString(width / 2, y_pos - 1.2 * cm, f"e-mail: {config.email_institucional}")
    
    # Primeira página
    y_atual = desenhar_cabecalho_pagina()
    y_atual = desenhar_cabecalho_tabela(y_atual)
    
    c.setFont("Helvetica", 9)
    
    for idx, pessoa_snap in enumerate(pessoas, 1):
        if y_atual - altura_linha < margem_inferior_tabela:
            c.showPage()
            y_atual = height - 1.5 * cm
            y_atual = desenhar_cabecalho_tabela(y_atual)
            c.setFont("Helvetica", 9)
        
        y_bottom = y_atual - altura_linha
        c.setFillColor(colors.black)
        c.rect(table_margin, y_bottom, total_table_width, altura_linha)
        
        current_x = table_margin
        for w in col_widths:
            c.line(current_x, y_atual, current_x, y_bottom)
            current_x += w
        c.line(current_x, y_atual, current_x, y_bottom)
        
        valor = float(pessoa_snap.valor_beneficio)
        valor_fmt = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        conta = memorando.conta_pagadora or ''
        
        y_text = y_atual - 0.5 * cm
        c.drawCentredString(col_x_starts[0] + col_widths[0]/2, y_text, str(idx))
        c.drawCentredString(col_x_starts[1] + col_widths[1]/2, y_text, pessoa_snap.nome_completo[:38])
        c.drawCentredString(col_x_starts[2] + col_widths[2]/2, y_text, beneficio.nome[:38])
        c.drawCentredString(col_x_starts[3] + col_widths[3]/2, y_text, valor_fmt)
        c.drawCentredString(col_x_starts[4] + col_widths[4]/2, y_text, str(conta)[:12])
        
        y_atual = y_bottom
    
    if y_atual - espaco_fechamento < 2 * cm:
        c.showPage()
        y_atual = height - 3 * cm
    
    desenhar_fechamento_assinatura(y_atual - 1.5 * cm)
    
    c.save()
    buffer.seek(0)
    return buffer

def gerar_pdf_beneficiarios(pessoas, beneficio_nome, status_label, 
                             total_pessoas, total_valor, total_ativos, total_espera, total_desligados):
    """Gera PDF do relatório de beneficiários"""
    from io import BytesIO
    from datetime import datetime
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), 
                           leftMargin=15*mm, rightMargin=15*mm, 
                           topMargin=15*mm, bottomMargin=15*mm)

    doc.title = 'Relatório de Beneficiários'
    styles = getSampleStyleSheet()
    elements = []
    
    # Brasão
    caminho_brasao = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static', 'images', 'brasao.jpg')
    if os.path.exists(caminho_brasao):
        img = Image(caminho_brasao, width=1.5*cm, height=1.5*cm)
        img.hAlign = 'CENTER'
        elements.append(img)
        elements.append(Spacer(1, 2*mm))
    
    # Cabeçalho institucional
    centro_style = ParagraphStyle('Centro', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, leading=14, fontName='Helvetica-Bold')
    elements.append(Paragraph('ESTADO DA PARAÍBA', centro_style))
    elements.append(Paragraph('PREFEITURA MUNICIPAL DE POCINHOS', centro_style))
    elements.append(Paragraph('SECRETARIA MUNICIPAL DE ASSISTÊNCIA SOCIAL', centro_style))
    elements.append(Spacer(1, 5*mm))
    
    # Título do relatório
    titulo_style = ParagraphStyle('Titulo', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=3)
    subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=8)
    
    elements.append(Paragraph('RELATÓRIO DE BENEFICIÁRIOS', titulo_style))
    elements.append(Paragraph(
        f'Benefício: {beneficio_nome} | Status: {status_label} | '
        f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        subtitulo_style
    ))
    elements.append(Spacer(1, 3*mm))
    
    # Tabela de dados
    header = ['Nº', 'Nome Completo', 'CPF', 'Benefício', 'Valor (R$)', 'Status', 'Bairro', 'Cadastro']
    data = [header]
    status_map = {'ativo': 'Ativo', 'em_espera': 'Em Espera', 'desligado': 'Desligado'}
    
    for idx, p in enumerate(pessoas, 1):
        cpf_numeros = ''.join(filter(str.isdigit, p.cpf))
        if len(cpf_numeros) == 11:
            cpf_display = f'{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}'
        else:
            cpf_display = p.cpf
        
        valor_fmt = f"{p.valor_beneficio:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        
        data.append([
            str(idx),
            Paragraph(p.nome_completo, ParagraphStyle('Cell', fontSize=7, leading=9)),
            cpf_display,
            Paragraph(p.beneficio.nome, ParagraphStyle('Cell', fontSize=7, leading=9)),
            valor_fmt,
            status_map.get(p.status, p.status),
            Paragraph(p.bairro, ParagraphStyle('Cell', fontSize=7, leading=9)),
            p.created_at.strftime('%d/%m/%Y'),
        ])
    
    col_widths = [25, 170, 75, 95, 55, 55, 80, 55]
    
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a7cff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (0, -1), 'CENTER'),
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 0), (5, -1), 'CENTER'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 8*mm))
    
    # Totalizadores fora da tabela
    valor_total_fmt = f"{total_valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    resumo_style = ParagraphStyle('Resumo', parent=styles['Normal'], fontSize=9, leading=14)
    
    elements.append(Paragraph(f'<b>Total de pessoas:</b> {total_pessoas}', resumo_style))
    elements.append(Paragraph(f'<b>Valor total mensal:</b> R$ {valor_total_fmt}', resumo_style))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph(f'<b>Ativos:</b> {total_ativos} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Em Espera:</b> {total_espera} &nbsp;&nbsp; | &nbsp;&nbsp; <b>Desligados:</b> {total_desligados}', resumo_style))
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="relatorio_beneficiarios.pdf"'
    return response


def gerar_excel_beneficiarios(pessoas, beneficio_nome, status_label,
                               total_pessoas, total_valor, total_ativos, total_espera, total_desligados):
    """Gera Excel do relatório de beneficiários"""
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Beneficiários'
    
    # Estilos
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4A7CFF', end_color='4A7CFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )
    
    # Título
    ws.merge_cells('A1:H1')
    ws['A1'] = 'Relatório de Beneficiários'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:H2')
    ws['A2'] = f'Benefício: {beneficio_nome} | Status: {status_label} | Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=9, color='808080')
    
    # Cabeçalho
    headers = ['Nº', 'Nome Completo', 'CPF', 'Benefício', 'Valor (R$)', 'Status', 'Bairro', 'Cadastro']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    
    # Dados
    status_map = {'ativo': 'Ativo', 'em_espera': 'Em Espera', 'desligado': 'Desligado'}
    
    for idx, p in enumerate(pessoas, 1):
        row = idx + 4
        cpf_numeros = ''.join(filter(str.isdigit, p.cpf))
        if len(cpf_numeros) == 11:
            cpf_display = f'{cpf_numeros[:3]}.{cpf_numeros[3:6]}.{cpf_numeros[6:9]}-{cpf_numeros[9:]}'
        else:
            cpf_display = p.cpf
        
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=p.nome_completo).border = border
        ws.cell(row=row, column=3, value=cpf_display).border = border
        ws.cell(row=row, column=4, value=p.beneficio.nome).border = border
        cell_valor = ws.cell(row=row, column=5, value=float(p.valor_beneficio))
        cell_valor.number_format = '#,##0.00'
        cell_valor.border = border
        ws.cell(row=row, column=6, value=status_map.get(p.status, p.status)).border = border
        ws.cell(row=row, column=7, value=p.bairro).border = border
        ws.cell(row=row, column=8, value=p.created_at.strftime('%d/%m/%Y')).border = border
        
        if idx % 2 == 0:
            for col in range(1, 9):
                ws.cell(row=row, column=col).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
    
    # Totalizadores
    total_row = len(pessoas) + 5
    total_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    total_font = Font(bold=True, size=10)
    
    ws.cell(row=total_row, column=2, value=f'Total: {total_pessoas} pessoas').font = total_font
    ws.cell(row=total_row, column=5, value=float(total_valor)).font = total_font
    ws.cell(row=total_row, column=5).number_format = '#,##0.00'
    ws.cell(row=total_row, column=6, value=f'Ativos: {total_ativos} | Em Espera: {total_espera} | Desligados: {total_desligados}').font = total_font
    
    for col in range(1, 9):
        ws.cell(row=total_row, column=col).fill = total_fill
        ws.cell(row=total_row, column=col).border = border
    
    # Larguras
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_beneficiarios.xlsx"'
    return response

def gerar_pdf_financeiro(dados):
    """Gera PDF do relatório financeiro"""
    from io import BytesIO
    from datetime import datetime
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm, cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from django.http import HttpResponse
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
                           leftMargin=15*mm, rightMargin=15*mm,
                           topMargin=15*mm, bottomMargin=15*mm)
    doc.title = 'Relatório Financeiro'
    
    styles = getSampleStyleSheet()
    elements = []
    
    # Brasão
    caminho_brasao = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'static', 'images', 'brasao.jpg')
    if os.path.exists(caminho_brasao):
        img = Image(caminho_brasao, width=1.5*cm, height=1.5*cm)
        img.hAlign = 'CENTER'
        elements.append(img)
        elements.append(Spacer(1, 2*mm))
    
    # Cabeçalho institucional
    centro_style = ParagraphStyle('Centro', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER, leading=14, fontName='Helvetica-Bold')
    elements.append(Paragraph('ESTADO DA PARAÍBA', centro_style))
    elements.append(Paragraph('PREFEITURA MUNICIPAL DE POCINHOS', centro_style))
    elements.append(Paragraph('SECRETARIA MUNICIPAL DE ASSISTÊNCIA SOCIAL', centro_style))
    elements.append(Spacer(1, 5*mm))
    
    titulo_style = ParagraphStyle('Titulo', parent=styles['Normal'], fontSize=12, alignment=TA_CENTER, fontName='Helvetica-Bold', spaceAfter=3)
    subtitulo_style = ParagraphStyle('Subtitulo', parent=styles['Normal'], fontSize=9, alignment=TA_CENTER, textColor=colors.grey, spaceAfter=8)
    secao_style = ParagraphStyle('Secao', parent=styles['Normal'], fontSize=10, fontName='Helvetica-Bold', spaceAfter=5, spaceBefore=10, alignment=TA_CENTER)
    resumo_style = ParagraphStyle('Resumo', parent=styles['Normal'], fontSize=9, leading=14)
    
    elements.append(Paragraph('RELATÓRIO FINANCEIRO', titulo_style))
    elements.append(Paragraph(
        f'Benefício: {dados["beneficio_label"]} | Status: {dados["status_label"]} | '
        f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}',
        subtitulo_style
    ))
    elements.append(Spacer(1, 5*mm))
    
    # ═══ RESUMO GERAL ═══
    rg = dados['resumo_geral']
    valor_fmt = f"{rg['total_valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    
    elements.append(Paragraph('RESUMO GERAL', secao_style))
    
    resumo_data = [
        ['Benefícios Ativos', str(rg['total_beneficios'])],
        ['Pessoas Ativas', str(rg['total_ativos'])],
        ['Valor Mensal Total', f'R$ {valor_fmt}'],
    ]
    
    resumo_table = Table(resumo_data, colWidths=[150, 150])
    resumo_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('TOPPADDING', (0, 0), (-1, -1), 4),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
    ]))
    elements.append(resumo_table)
    elements.append(Spacer(1, 8*mm))
    
    # ═══ DETALHAMENTO POR BENEFÍCIO ═══
    elements.append(Paragraph('DETALHAMENTO POR BENEFÍCIO', secao_style))
    
    ben_header = ['Descrição', 'Benefício', 'Ativos', 'Espera', 'Deslig.', 'Valor Mensal', '%']
    ben_data = [ben_header]
    
    cell_style = ParagraphStyle('CellFin', fontSize=7, leading=9)
    
    for b in dados['det_beneficios']:
        valor_b = f"{b['valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        ben_data.append([
            Paragraph(b['descricao'], cell_style),
            Paragraph(b['nome_oficial'], cell_style),
            str(b['ativos']),
            str(b['espera']),
            str(b['desligados']),
            f'R$ {valor_b}',
            f"{b['percentual']:.1f}%",
        ])
    
    # Linha total
    total_valor_fmt = f"{dados['total_valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    ben_data.append([
        Paragraph('<b>TOTAL</b>', cell_style),
        '',
        str(dados['total_ativos']),
        str(dados['total_espera']),
        str(dados['total_desligados']),
        f'R$ {total_valor_fmt}',
        '100%',
    ])
    
    ben_widths = [120, 120, 40, 40, 40, 85, 35]
    ben_table = Table(ben_data, colWidths=ben_widths, repeatRows=1)
    ben_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a7cff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (2, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (5, 0), (5, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(ben_table)
    elements.append(Spacer(1, 8*mm))
    
    # ═══ DETALHAMENTO POR FAIXA DE VALOR ═══
    elements.append(Paragraph('DETALHAMENTO POR FAIXA DE VALOR', secao_style))
    
    faixa_header = ['Faixa', 'Pessoas', 'Valor Total', '%']
    faixa_data = [faixa_header]
    
    for f in dados['det_faixas']:
        valor_f = f"{f['valor']:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        faixa_data.append([
            f['faixa'],
            str(f['pessoas']),
            f'R$ {valor_f}',
            f"{f['percentual']:.1f}%",
        ])
    
    # Total faixas
    total_faixa_pessoas = sum(f['pessoas'] for f in dados['det_faixas'])
    faixa_data.append([
        'TOTAL',
        str(total_faixa_pessoas),
        f'R$ {total_valor_fmt}',
        '100%',
    ])
    
    faixa_widths = [120, 60, 100, 50]
    faixa_table = Table(faixa_data, colWidths=faixa_widths, repeatRows=1)
    faixa_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4a7cff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (2, 0), (2, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#f8f9fa')]),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#e9ecef')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    elements.append(faixa_table)
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="relatorio_financeiro.pdf"'
    return response


def gerar_excel_financeiro(dados):
    """Gera Excel do relatório financeiro"""
    from io import BytesIO
    from datetime import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from django.http import HttpResponse
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Financeiro'
    
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='4A7CFF', end_color='4A7CFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin', color='DEE2E6'),
        right=Side(style='thin', color='DEE2E6'),
        top=Side(style='thin', color='DEE2E6'),
        bottom=Side(style='thin', color='DEE2E6'),
    )
    total_fill = PatternFill(start_color='E9ECEF', end_color='E9ECEF', fill_type='solid')
    total_font = Font(bold=True, size=10)
    secao_font = Font(bold=True, size=11)
    
    # Título
    ws.merge_cells('A1:G1')
    ws['A1'] = 'Relatório Financeiro'
    ws['A1'].font = Font(bold=True, size=14)
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f'Benefício: {dados["beneficio_label"]} | Status: {dados["status_label"]} | Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
    ws['A2'].font = Font(size=9, color='808080')
    
    # ═══ RESUMO GERAL ═══
    row = 4
    ws.cell(row=row, column=1, value='RESUMO GERAL').font = secao_font
    row = 5
    
    rg = dados['resumo_geral']
    valor_total = float(rg['total_valor'])
    
    resumo_items = [
        ('Benefícios Ativos', rg['total_beneficios']),
        ('Pessoas Ativas', rg['total_ativos']),
        ('Valor Mensal Total', valor_total),
    ]
    
    for label, valor in resumo_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=1).border = border
        cell = ws.cell(row=row, column=2, value=valor)
        cell.border = border
        if isinstance(valor, float):
            cell.number_format = 'R$ #,##0.00'
        row += 1
    
    # ═══ DETALHAMENTO POR BENEFÍCIO ═══
    row += 1
    ws.cell(row=row, column=1, value='DETALHAMENTO POR BENEFÍCIO').font = secao_font
    row += 1
    
    ben_headers = ['Descrição', 'Nome Oficial', 'Ativos', 'Espera', 'Deslig.', 'Valor Mensal', '%']
    for col, h in enumerate(ben_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    row += 1
    
    for idx, b in enumerate(dados['det_beneficios']):
        ws.cell(row=row, column=1, value=b['descricao']).border = border
        ws.cell(row=row, column=2, value=b['nome_oficial']).border = border
        ws.cell(row=row, column=3, value=b['ativos']).border = border
        ws.cell(row=row, column=4, value=b['espera']).border = border
        ws.cell(row=row, column=5, value=b['desligados']).border = border
        cell_v = ws.cell(row=row, column=6, value=float(b['valor']))
        cell_v.number_format = '#,##0.00'
        cell_v.border = border
        ws.cell(row=row, column=7, value=f"{b['percentual']:.1f}%").border = border
        
        if idx % 2 == 1:
            for c in range(1, 8):
                ws.cell(row=row, column=c).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        row += 1
    
    # Total benefícios
    ws.cell(row=row, column=1, value='TOTAL').font = total_font
    ws.cell(row=row, column=3, value=dados['total_ativos']).font = total_font
    ws.cell(row=row, column=4, value=dados['total_espera']).font = total_font
    ws.cell(row=row, column=5, value=dados['total_desligados']).font = total_font
    cell_tv = ws.cell(row=row, column=6, value=float(dados['total_valor']))
    cell_tv.number_format = '#,##0.00'
    cell_tv.font = total_font
    ws.cell(row=row, column=7, value='100%').font = total_font
    for c in range(1, 8):
        ws.cell(row=row, column=c).fill = total_fill
        ws.cell(row=row, column=c).border = border
    
    # ═══ DETALHAMENTO POR FAIXA DE VALOR ═══
    row += 2
    ws.cell(row=row, column=1, value='DETALHAMENTO POR FAIXA DE VALOR').font = secao_font
    row += 1
    
    faixa_headers = ['Faixa', 'Pessoas', 'Valor Total', '%']
    for col, h in enumerate(faixa_headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
    row += 1
    
    for idx, f in enumerate(dados['det_faixas']):
        ws.cell(row=row, column=1, value=f['faixa']).border = border
        ws.cell(row=row, column=2, value=f['pessoas']).border = border
        cell_fv = ws.cell(row=row, column=3, value=float(f['valor']))
        cell_fv.number_format = '#,##0.00'
        cell_fv.border = border
        ws.cell(row=row, column=4, value=f"{f['percentual']:.1f}%").border = border
        
        if idx % 2 == 1:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = PatternFill(start_color='F8F9FA', end_color='F8F9FA', fill_type='solid')
        row += 1
    
    # Total faixas
    total_faixa_pessoas = sum(f['pessoas'] for f in dados['det_faixas'])
    ws.cell(row=row, column=1, value='TOTAL').font = total_font
    ws.cell(row=row, column=2, value=total_faixa_pessoas).font = total_font
    cell_tfv = ws.cell(row=row, column=3, value=float(dados['total_valor']))
    cell_tfv.number_format = '#,##0.00'
    cell_tfv.font = total_font
    ws.cell(row=row, column=4, value='100%').font = total_font
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = total_fill
        ws.cell(row=row, column=c).border = border
    
    # Larguras
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 16
    ws.column_dimensions['G'].width = 8
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    response = HttpResponse(buffer, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="relatorio_financeiro.xlsx"'
    return response